import pyrebase
import face_recognition
from PIL import Image, ImageDraw
import os,sys
import numpy as np
import openpyxl
import smtplib


config = {
	  "apiKey": "AIzaSyDZFaXexrptFx5dGnLnG04SApyS9N-Vb_c",
	  "authDomain": "faceattendance-89125.firebaseapp.com",
	  "databaseURL": "https://faceattendance-89125.firebaseio.com",
	  "projectId": "faceattendance-89125",
	  "storageBucket": "faceattendance-89125.appspot.com",
	  "messagingSenderId": "1088230635859"
	 

	}
firebase = pyrebase.initialize_app(config)
storage = firebase.storage()

attendance_doc=openpyxl.load_workbook('attendance.xlsx')
sheet = attendance_doc.get_sheet_by_name('Sheet1')

no_img=True

#downoading firebase class image
def download():
	storage.child("faceimage.jpg").download("./unknown/current.jpg")
	if os.path.exists("./unknown/current.jpg"):
		if os.stat('./unknown/prev.jpg').st_size!=os.stat('./unknown/current.jpg').st_size:
			return 1
	return 0


while True:
	no_img=True
	while no_img==True:
		i=download()
		if i==0:
			print("NO IMAGE UPLOADED")
			continue
		print("IMAGE FOUND")
		print("PROCESSING...")
		no_img=False
	# This is an example of running face recognition on a single image
	# and drawing a box around each person that was identified.

	for i in range(3,7):
		sheet.cell(row=i, column=2).value="Absent"
	
	pr=""

	known_face_encodings=[]
	known_face_names=[]

	for files in os.listdir("./know"):	
		image = face_recognition.load_image_file("./know/"+files)
		face_encoding = face_recognition.face_encodings(image)[0]
		known_face_encodings.append(face_encoding)
		known_face_names.append(files[:-4])
	# Load an image with an unknown face
	unknown_image = face_recognition.load_image_file("./unknown/current.jpg")

	# Find all the faces and face encodings in the unknown image
	face_locations = face_recognition.face_locations(unknown_image)
	face_encodings = face_recognition.face_encodings(unknown_image, face_locations)

	# Convert the image to a PIL-format image so that we can draw on top of it with the Pillow library
	# See http://pillow.readthedocs.io/ for more about PIL/Pillow
	pil_image = Image.fromarray(unknown_image)
	# Create a Pillow ImageDraw Draw instance to draw with
	draw = ImageDraw.Draw(pil_image)

	# Loop through each face found in the unknown image
	for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
	    # See if the face is a match for the known face(s)
		matches = face_recognition.compare_faces(known_face_encodings, face_encoding , tolerance=0.45)

		name = "Unknown"

	    # If a match was found in known_face_encodings, just use the first one.
	    # if True in matches:
	    #     first_match_index = matches.index(True)
	    #     name = known_face_names[first_match_index]

	    # Or instead, use the known face with the smallest distance to the new face
		face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
		best_match_index = np.argmin(face_distances)
		if matches[best_match_index]:
			name = known_face_names[best_match_index]

		for i in range(3,7):
			if sheet.cell(row=i, column=1).value == name:
				sheet.cell(row=i, column=2).value="Present"
				pr=pr+"\n"+name
			

	    # Draw a label with a name below the face



	# Remove the drawing library from memory as per the Pillow docs
	del draw
	# Python code to illustrate Sending mail from  
	# your Gmail account 
	  
	# creates SMTP session 
	s = smtplib.SMTP('smtp.gmail.com', 587) 
	  
	# start TLS for security 
	s.starttls() 
	  
	# Authentication 
	s.login("facerecognitionforpbl@gmail.com", "face4you") 
	  
	# message to be sent 
	message = "\n\nPresent:-  " +pr
	  
	li=["kvnl.hritik@gmail.com","vikramborana9603@gmail.com"]

	for i in range(len(li)): 
	# sending the mail 
		s.sendmail("facerecognitionforpbl@gmail.com",li[i], message) 
	  
	# terminating the session 
	s.quit() 

	

	attendance_doc.save('attendance.xlsx')
	os.remove("unknown/prev.jpg")
	os.rename(r'unknown/current.jpg',r'unknown/prev.jpg')
	print("\n\nNEW EXECUTION\n")
	continue
